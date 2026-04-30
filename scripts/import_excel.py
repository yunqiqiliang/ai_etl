"""导入猫砂竞品 Excel 数据到 ClickZetta Lakehouse

用法:
    PYTHONPATH=. python scripts/import_excel.py /path/to/猫砂数据.xlsx

前提:
    - .env 中配置了 CLICKZETTA_USERNAME / CLICKZETTA_PASSWORD
    - config.yaml 中配置了 clickzetta 连接信息
    - pip install openpyxl
"""
import sys
import logging

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

from pathlib import Path
from ai_etl.config import Config
from ai_etl.lakehouse import LakehouseClient

# ── Sheet → 表名映射 ─────────────────────────────────────
SHEET_TABLE_MAP = {
    "福丸竞品主数据": "product_master",
    "猫砂抖音": "sales_douyin",
    "猫砂天猫": "sales_tmall",
    "猫砂天猫超市": "sales_tmall_chaoshi",
    "猫砂京东自营": "sales_jd_self",
    "猫砂京东pop": "sales_jd_pop",
    "类目表": "category_mapping",
}

TABLE_COMMENTS = {
    "product_master": "福丸品牌SKU主数据(人工标注标杆数据,46条)",
    "sales_douyin": "抖音猫砂销售数据(精确值)",
    "sales_tmall": "天猫猫砂销售数据(件单价+区间销量)",
    "sales_tmall_chaoshi": "天猫超市猫砂销售数据(指数值,非真实金额)",
    "sales_jd_self": "京东自营猫砂销售排行(区间值,如￥50万~￥75万)",
    "sales_jd_pop": "京东POP猫砂销售排行(区间值)",
    "category_mapping": "猫砂四级类目编码映射表(16个类目)",
}


def import_excel(excel_path: str, schema: str = "cat_litter"):
    """导入 Excel 到 Lakehouse。"""
    try:
        import openpyxl
    except ImportError:
        print("请先安装 openpyxl: pip install openpyxl")
        sys.exit(1)

    from clickzetta.zettapark.types import StringType, StructField, StructType

    excel_path = Path(excel_path)
    if not excel_path.exists():
        print(f"文件不存在: {excel_path}")
        sys.exit(1)

    lh = LakehouseClient(config=Config())

    # 创建 schema
    lh.session.sql(
        f"CREATE SCHEMA IF NOT EXISTS {schema} COMMENT '猫砂竞品分析'"
    ).collect()
    logger.info("Schema %s 已就绪", schema)

    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
    logger.info("Excel 文件: %s (%d sheets)", excel_path.name, len(wb.sheetnames))

    for sheet_name in wb.sheetnames:
        table_name = SHEET_TABLE_MAP.get(sheet_name)
        if not table_name:
            logger.warning("跳过未映射的 sheet: %s", sheet_name)
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            logger.warning("Sheet %s 无数据行，跳过", sheet_name)
            continue

        # 第一行是列名
        raw_headers = rows[0]
        headers = []
        for h in raw_headers:
            h = str(h).strip() if h else ""
            if not h:
                h = f"col_{len(headers)}"
            headers.append(h)

        # 数据行（跳过全空行）
        data_rows = []
        for row in rows[1:]:
            values = [str(v).strip() if v is not None else "" for v in row]
            if any(v for v in values):
                # 补齐列数
                while len(values) < len(headers):
                    values.append("")
                data_rows.append(values[: len(headers)])

        if not data_rows:
            logger.warning("Sheet %s 过滤后无数据，跳过", sheet_name)
            continue

        # 全部用 STRING 类型
        struct_fields = [StructField(h, StringType()) for h in headers]
        write_schema = StructType(struct_fields)

        full_table = f"{schema}.{table_name}"
        logger.info(
            "导入 %s → %s (%d 行, %d 列)",
            sheet_name, full_table, len(data_rows), len(headers),
        )

        # 先删除旧表
        lh.session.sql(f"DROP TABLE IF EXISTS {full_table}").collect()

        df = lh.session.create_dataframe(data_rows, schema=write_schema)
        df.write.save_as_table(full_table, mode="append")

        # 添加表注释
        comment = TABLE_COMMENTS.get(table_name, sheet_name)
        try:
            lh.session.sql(
                f"ALTER TABLE {full_table} SET COMMENT '{comment}'"
            ).collect()
        except Exception:
            pass

        logger.info("  ✅ %s: %d 行", full_table, len(data_rows))

    wb.close()

    # 验证
    print("\n--- 导入结果 ---")
    for table_name in SHEET_TABLE_MAP.values():
        full_table = f"{schema}.{table_name}"
        try:
            cnt = lh.session.sql(
                f"SELECT COUNT(*) AS c FROM {full_table}"
            ).collect()[0]["c"]
            print(f"  {full_table}: {cnt} 行")
        except Exception:
            print(f"  {full_table}: 不存在")

    lh.close()
    print("\n导入完成!")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: PYTHONPATH=. python scripts/import_excel.py <excel文件路径>")
        print("示例: PYTHONPATH=. python scripts/import_excel.py ~/Documents/猫砂数据.xlsx")
        sys.exit(1)
    import_excel(sys.argv[1])
